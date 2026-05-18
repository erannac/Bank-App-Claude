from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DARK_BLUE = "1E3A5F"
WHITE = "FFFFFF"
ROW_ALT = "EEF2F7"
RED_LIGHT = "FFCCCC"
ORANGE_LIGHT = "FFE5CC"
YELLOW_LIGHT = "FFFACC"
GREEN_LIGHT = "CCFFCC"
INCOME_BG = "E8F5E9"
EXPENSE_BG = "FFEBEE"
FEES_HEADER = "4A1A6B"

SEVERITY_BG = {"high": RED_LIGHT, "medium": ORANGE_LIGHT, "low": YELLOW_LIGHT}
SEVERITY_LABEL = {"high": "גבוהה", "medium": "בינונית", "low": "נמוכה"}

FEE_KEYWORDS = ("עמלה", "עמל'", "עמל.", "דמי", "ריבית", "עמלות", "ועמלה", "קנס", "היטל")


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_style(cell, bg: str = DARK_BLUE, fg: str = WHITE) -> None:
    cell.font = Font(bold=True, color=fg, size=11, name="Arial")
    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _auto_width(ws) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _title_row(ws, text: str, span: str, row: int = 1, bg: str = DARK_BLUE) -> None:
    ws.merge_cells(span)
    cell = ws[span.split(":")[0]]
    cell.value = text
    cell.font = Font(bold=True, size=14, color=WHITE, name="Arial")
    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32


def _parse_month_key(m: str) -> tuple:
    """Return (year, month) for sorting MM/YYYY strings."""
    try:
        parts = m.split("/")
        return (int(parts[1]), int(parts[0]))
    except Exception:
        return (0, 0)


def _is_fee(tx: dict) -> bool:
    cat = (tx.get("category") or "").lower()
    desc = (tx.get("description") or "").lower()
    if "עמלות" in cat:
        return True
    return any(kw in desc for kw in FEE_KEYWORDS)


def generate_report(analysis: dict, output_path: str) -> None:
    wb = Workbook()
    _sheet_summary(wb, analysis)
    _sheet_categories(wb, analysis)
    _sheet_transactions(wb, analysis)
    _sheet_fees(wb, analysis)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(output_path)


# ── Sheet 1: Executive summary ────────────────────────────────────────────────

def _sheet_summary(wb: Workbook, analysis: dict) -> None:
    ws = wb.create_sheet("סיכום מנהלים", 0)
    ws.sheet_view.rightToLeft = True

    _title_row(ws, "דוח ניתוח פיננסי – סיכום מנהלים", "A1:F1", 1)

    ws["A3"] = "סיכום ממצאים:"
    ws["A3"].font = Font(bold=True, size=11, name="Arial")
    ws.merge_cells("A4:F4")
    ws["A4"] = analysis.get("summary", "")
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[4].height = 70

    suspicious = analysis.get("suspicious", [])
    ws["A6"] = f"תנועות חשודות ({len(suspicious)} ממצאים):"
    ws["A6"].font = Font(bold=True, size=11, color="C0392B", name="Arial")

    for col, h in enumerate(["חומרה", "סוג", "תיאור", "סכום (₪)", "תאריך"], 1):
        _header_style(ws.cell(row=7, column=col, value=h))

    for r, item in enumerate(suspicious, 8):
        sev = item.get("severity", "low")
        bg = SEVERITY_BG.get(sev, WHITE)
        amount = item.get("amount", "")
        if isinstance(amount, (int, float)) and amount > 0:
            amount = -abs(amount)
        for col, v in enumerate(
            [SEVERITY_LABEL.get(sev, sev), item.get("type", ""), item.get("description", ""), amount, item.get("date", "")],
            1,
        ):
            cell = ws.cell(row=r, column=col, value=v)
            cell.fill = PatternFill(fill_type="solid", fgColor=bg)
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    _auto_width(ws)


# ── Sheet 2: Categories — pivot (category × month) ───────────────────────────

def _sheet_categories(wb: Workbook, analysis: dict) -> None:
    ws = wb.create_sheet("הוצאות לפי קטגוריה", 1)
    ws.sheet_view.rightToLeft = True

    categories = analysis.get("categories", {})

    # Collect and sort all months
    all_months: set[str] = set()
    for cat_data in categories.values():
        all_months.update(cat_data.get("months", {}).keys())
    sorted_months = sorted(all_months, key=_parse_month_key)

    num_cols = 1 + len(sorted_months) + 1  # category | months... | total
    last_col = get_column_letter(num_cols)
    _title_row(ws, "הוצאות לפי קטגוריה — לפי חודש", f"A1:{last_col}1", 1)

    # Header row
    headers = ["קטגוריה"] + sorted_months + ['סה"כ']
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(row=3, column=col, value=h))

    # Data rows — one row per category
    for r, (cat_name, cat_data) in enumerate(categories.items(), 4):
        months_data = cat_data.get("months", {})
        is_income = "הכנסה" in cat_name

        bg = ROW_ALT if r % 2 == 0 else WHITE
        amount_color = "1A56DB" if is_income else "000000"  # blue for income, black for expense

        # Category name
        name_cell = ws.cell(row=r, column=1, value=cat_name)
        name_cell.fill = PatternFill(fill_type="solid", fgColor=bg)
        name_cell.font = Font(bold=True, name="Arial", size=10)
        name_cell.border = _thin_border()
        name_cell.alignment = Alignment(horizontal="right", vertical="center")

        row_total = 0.0
        for col_idx, month in enumerate(sorted_months, 2):
            raw_val = months_data.get(month, None)
            if raw_val is not None:
                try:
                    amount = float(raw_val)
                    if not is_income and amount > 0:
                        amount = -amount
                    row_total += amount
                except (ValueError, TypeError):
                    amount = raw_val
            else:
                amount = ""

            cell = ws.cell(row=r, column=col_idx, value=amount)
            cell.fill = PatternFill(fill_type="solid", fgColor=bg)
            cell.font = Font(name="Arial", size=10, color=amount_color)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(amount, float):
                cell.number_format = '#,##0.00;[Red]-#,##0.00'

        # Total column
        total_cell = ws.cell(row=r, column=num_cols, value=round(row_total, 2) if row_total else "")
        total_cell.fill = PatternFill(fill_type="solid", fgColor=bg)
        total_cell.font = Font(bold=True, name="Arial", size=10, color=amount_color)
        total_cell.border = _thin_border()
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
        if isinstance(row_total, float):
            total_cell.number_format = '#,##0.00;[Red]-#,##0.00'

    _auto_width(ws)
    # Freeze category column
    ws.freeze_panes = "B4"


# ── Sheet 3: All transactions ─────────────────────────────────────────────────

def _sheet_transactions(wb: Workbook, analysis: dict) -> None:
    ws = wb.create_sheet("כל התנועות", 2)
    ws.sheet_view.rightToLeft = True

    _title_row(ws, "רשימת כל התנועות", "A1:D1", 1)

    for col, h in enumerate(["תאריך", "תיאור", "סכום (₪)", "קטגוריה"], 1):
        _header_style(ws.cell(row=3, column=col, value=h))

    for r, tx in enumerate(analysis.get("transactions", []), 4):
        cat = tx.get("category", "")
        is_income = "הכנסה" in cat
        raw_amount = tx.get("amount", "")

        try:
            amount = float(raw_amount)
            if not is_income and amount > 0:
                amount = -amount
        except (ValueError, TypeError):
            amount = raw_amount

        bg = ROW_ALT if r % 2 == 0 else WHITE

        for col, v in enumerate([tx.get("date", ""), tx.get("description", ""), amount, cat], 1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.fill = PatternFill(fill_type="solid", fgColor=bg)
            cell.border = _thin_border()
            if col == 3 and isinstance(v, float):
                cell.number_format = '#,##0.00;[Red]-#,##0.00'
                cell.font = Font(name="Arial", size=10, color="1A56DB" if is_income else "000000")

    _auto_width(ws)


# ── Sheet 4: Fees analysis ────────────────────────────────────────────────────

def _sheet_fees(wb: Workbook, analysis: dict) -> None:
    ws = wb.create_sheet("ניתוח עמלות", 3)
    ws.sheet_view.rightToLeft = True

    _title_row(ws, "ניתוח עמלות בנק וכרטיסי אשראי", "A1:E1", 1, bg=FEES_HEADER)

    fee_txs = [tx for tx in analysis.get("transactions", []) if _is_fee(tx)]

    # Summary box
    total_fees = 0.0
    for tx in fee_txs:
        try:
            total_fees += abs(float(tx.get("amount", 0)))
        except (ValueError, TypeError):
            pass

    ws["A3"] = "סה\"כ עמלות שזוהו:"
    ws["A3"].font = Font(bold=True, size=11, name="Arial")
    ws["B3"] = -round(total_fees, 2)
    ws["B3"].font = Font(bold=True, size=12, color="C0392B", name="Arial")
    ws["B3"].number_format = '#,##0.00;[Red]-#,##0.00'

    ws["A4"] = f"מספר עמלות: {len(fee_txs)}"
    ws["A4"].font = Font(size=10, name="Arial", color="555555")

    # Table header
    for col, h in enumerate(["תאריך", "תיאור", "סכום (₪)", "קטגוריה", "הערה"], 1):
        _header_style(ws.cell(row=6, column=col, value=h), bg=FEES_HEADER)

    if not fee_txs:
        ws.cell(row=7, column=1, value="לא זוהו עמלות בנתונים שסופקו")
        ws.merge_cells("A7:E7")
    else:
        for r, tx in enumerate(fee_txs, 7):
            try:
                amount = -abs(float(tx.get("amount", 0)))
            except (ValueError, TypeError):
                amount = tx.get("amount", "")

            bg = ROW_ALT if r % 2 == 0 else WHITE
            note = "⚠️ חשוד" if any(
                s.get("description", "") and tx.get("description", "") and
                tx["description"][:10] in s["description"]
                for s in analysis.get("suspicious", [])
            ) else ""

            for col, v in enumerate([tx.get("date", ""), tx.get("description", ""), amount, tx.get("category", ""), note], 1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.fill = PatternFill(fill_type="solid", fgColor=bg)
                cell.border = _thin_border()
                if col == 3 and isinstance(v, float):
                    cell.number_format = '#,##0.00;[Red]-#,##0.00'

    _auto_width(ws)
